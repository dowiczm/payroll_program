import pandas as pd
from datetime import datetime
import openpyxl

import warnings
warnings.filterwarnings("ignore")

from b_auto_cleaner import auto_cleaner
from c_to_float import to_float
from d_clean_names import clean_employee_names
from create_file import create_file

import os

def master_payroll(service_path, product_path, info_path, destination_path, date):
    
    raw_service = pd.read_csv(f'{service_path[-1]}').drop(columns=['Unnamed: 7', 'Unnamed: 10'])
    raw_product = pd.read_csv(f'{product_path[-1]}')[['Commissions and Gratuity', 'Unnamed: 1',
                                                    'Unnamed: 2', 'Unnamed: 3', 'Unnamed: 4', 'Unnamed: 5',
                                                    'Unnamed: 6', 'Unnamed: 7', 'Unnamed: 8', 'Unnamed: 9',
                                                    'Unnamed: 10', 'Unnamed: 11', 'Unnamed: 12', 'Unnamed: 13',
                                                    'Unnamed: 14']]

    raw_product.drop(columns=['Unnamed: 7', 'Unnamed: 10'], inplace=True)


    info = pd.read_excel(f'{info_path[-1]}', sheet_name='Info')
    destination = destination_path[-1] + '\\'


    
    services = auto_cleaner(raw_service)
    products = auto_cleaner(raw_product)
    
    to_float(services)
    to_float(products)
    
    clean_employee_names(services)
    clean_employee_names(products)
    
    medical_professionals = []
    for name, typ in zip(info['Name'], info['Type']):
        if typ == 'Ae':
            medical_professionals.append(name)

    massage_therapists = []
    for name, typ in zip(info['Name'], info['Type']):
        if typ == 'LMT':
            massage_therapists.append(name)

    ests = []
    for name, typ in zip(info['Name'], info['Type']):
        if typ == 'Med':
            ests.append(name)

    front_desk = []
    for name, typ in zip(info['Name'], info['Type']):
        if typ == 'FNT':
            front_desk.append(name)

    management = []
    for name, typ in zip(info['Name'], info['Type']):
        if typ == 'Mgt':
            management.append(name)

    all_employees = [medical_professionals, massage_therapists, ests, front_desk, management]


    #Create File
    create_file(services, products, medical_professionals, massage_therapists, ests, management, front_desk, all_employees, destination)
    

    
    for n, name in enumerate(medical_professionals):
        med_dict = {
        'Amie Barrett' : 30,
        'Andrea Ward' : 30,
        'Jennifer Efstathiou' : 40,
        'Stephanie McDonagh' : 30
    }


        df = services.loc[services['Employee'] == name]
        wb = openpyxl.load_workbook(f'{destination}Master Payroll {date}.xlsx')
        worksheet = wb[name]

        length = len(df)

        #charge cell
        charge_cell = f'E{length + 4}'
        charge_formula = f'=SUM(E2:E{length+1})'
        worksheet[charge_cell] = charge_formula

        charge_title = f'E{length+3}'
        worksheet[charge_title] = 'Charge Total'

        #Rate
        rate_title = f'E{length+6}'
        worksheet[rate_title] = 'Rate'

        rate_formula = f'E{length+7}'
        worksheet[rate_formula] = med_dict[name]

        #net sum
        net_cell = f'G{length + 4}'
        net_formula = f'=SUM(G2:G{length+1})'
        worksheet[net_cell] = net_formula

        net_title = f'G{length+3}'
        worksheet[net_title] = 'Net Total'

        #less 10%
        less_10_cell = f'H{length+4}'
        less_10_formula = f'=(E{length+4})*0.1'
        worksheet[less_10_cell] = less_10_formula

        less_title = f'H{length+3}'
        worksheet[less_title] = 'Less 10%'

        #service income
        service_cell = f'I{length+4}'
        service_formula = f'=G{length+4}-H{length+4}'
        worksheet[service_cell] = service_formula

        service_title = f'I{length+3}'
        worksheet[service_title] = 'Gross Total'

        #realized service income
        income_cell = f'K{length+4}'
        income_formula = f'=I{length+4}*(E{length+7}/100)'
        worksheet[income_cell] = income_formula

        income_title = f'K{length+3}'
        worksheet[income_title] = 'Provider Total'

        #retail commission
        retail_commission = products.loc[products['Employee'] == name, 'Seller'].sum()
        retail_title_cell = f'K{length+6}'
        retail_cell = f'K{length+7}'

        worksheet[retail_title_cell] = 'Retail Commission'
        worksheet[retail_cell] = retail_commission

        #gratuity
        gratuity = products.loc[products['Employee'] == name, 'Gratuity'].sum()
        gratuity_title_cell = f'K{length+9}'
        gratuity_cell = f'K{length+10}'

        worksheet[gratuity_title_cell] = 'Gratuity'
        worksheet[gratuity_cell] = gratuity

        #income total
        income_title_cell = f'N{length+3}'
        worksheet[income_title_cell] = 'Total Income'

        income_cell = f'N{length+4}'
        income_formula = f'=SUM(K{length+4},K{length+7},K{length+10})'
        worksheet[income_cell] = income_formula


        wb.save(f'{destination}Master Payroll {date}.xlsx')


        #Cover Sheet
        wb = openpyxl.load_workbook(f'{destination}Master Payroll {date}.xlsx')
        worksheet = wb['Cover Sheet']
        row = str(n+2)

        reg_pay_cell = f'D{row}'
        worksheet[reg_pay_cell] = f"='{name}'!K{length+4}"

        gratuity_cell = f'G{row}'
        worksheet[gratuity_cell] = f"='{name}'!K{length+7}"

        commission_cell = f'H{row}'
        worksheet[commission_cell] = f"='{name}'!K{length+10}"

        total_cell = f'K{row}'
        worksheet[total_cell] = f'=SUM({reg_pay_cell},{gratuity_cell},{commission_cell},E{row},F{row})'


        wb.save(f'{destination}Master Payroll {date}.xlsx')





    #Massage Therapist Formulas
    massage_therapists = {
        'Debra Gottlieb' : 50,
        'Kim Moffat' : 65,
        'Kimberly Kristoff' : 55,
        'Rebecca Landry' : 55,
        'Robin Paolozzi' : 50,
        'Sean Flynn' : 50
    }

    for n, name in enumerate(massage_therapists):
        wb = openpyxl.load_workbook(f'{destination}Master Payroll {date}.xlsx')
        worksheet = wb[name]

        df = services.loc[services['Employee'] == name]

        length = len(df)

        #quantity total
        qty_title_cell = f'D{length+3}'
        worksheet[qty_title_cell] = 'Total Hours'

        qty_formula_cell = f'D{length+4}'
        worksheet[qty_formula_cell] = f'=SUM(D2:D{length+1})'

        #Rate
        rate_title = f'D{length+6}'
        worksheet[rate_title] = 'Rate'

        rate_formula = f'D{length+7}'
        worksheet[rate_formula] = massage_therapists[name]

        #service total
        service_title_cell = f'K{length+3}'
        worksheet[service_title_cell] = f'Provider Total'

        service_formula_cell = f'K{length+4}'
        worksheet[service_formula_cell] = f'={qty_formula_cell}*{rate_formula}'

        #commission total
        commission_title_cell = f'K{length+6}'
        worksheet[commission_title_cell] = f'Retail Commission'

        commission_cell = f'K{length+7}'
        worksheet[commission_cell] = products.loc[products['Employee'] == name]['Seller'].sum()

        #gratuity total
        gratuity_title_cell = f'K{length+9}'
        worksheet[gratuity_title_cell] = 'Gratuity Total'

        gratuity_cell = f'K{length+10}'
        worksheet[gratuity_cell] = products.loc[products['Employee'] == name]['Gratuity'].sum()

        #grand total
        total_cell_title = f'N{length+3}'
        worksheet[total_cell_title] = 'Total Income'

        total_cell = f'N{length+4}'
        worksheet[total_cell] = f'=SUM(K{length+4},K{length+7},K{length+10})'

        wb.save(f'{destination}Master Payroll {date}.xlsx')

        #Cover Sheet
        row = str(len(medical_professionals) + n + 3)
        wb = openpyxl.load_workbook(f'{destination}Master Payroll {date}.xlsx')
        worksheet = wb['Cover Sheet']

        reg_pay_cell = f'D{row}'
        worksheet[reg_pay_cell] = f"='{name}'!K{length+4}"

        retail_commission = f'G{row}'
        worksheet[retail_commission] = f"='{name}'!K{length+7}"

        gratuity_cell = f'H{row}'
        worksheet[gratuity_cell] = f"='{name}'!K{length+10}"

        total_cell = f'K{row}'
        worksheet[total_cell] = f'=SUM(D{row},E{row},F{row},G{row},H{row})'

        wb.save(f'{destination}Master Payroll {date}.xlsx')



    #Esth Formulas
    for n, name in enumerate(ests):
        wb = openpyxl.load_workbook(f'{destination}Master Payroll {date}.xlsx')
        worksheet = wb[name]

        df = services.loc[services['Employee'] == name]
        length = len(df)

        #provider sum
        provider_total_title = f'K{length+3}'
        worksheet[provider_total_title] = 'Provider Sum'

        provider_total = f'K{length+4}'
        worksheet[provider_total] = f'=SUM(K2:K{length+1})'

        #product commission
        commission_title = f'K{length+6}'
        worksheet[commission_title] = 'Product Commission'

        commission_value = f'K{length+7}'
        worksheet[commission_value] = products.loc[products['Employee'] == name]['Seller'].sum()

        #gratuity
        gratuity_title = f'K{length+9}'
        worksheet[gratuity_title] = 'Gratuity'

        gratuity_value = f'K{length+10}'
        worksheet[gratuity_value] = products.loc[products['Employee'] == name]['Gratuity'].sum()

        #income total
        worksheet[f'N{length+3}'] = 'Total Income'
        worksheet[f'N{length+4}'] = f'=SUM({provider_total},{commission_value},{gratuity_value})'

        if name == 'Tom Bondarchuk APRN':
            worksheet[f'L{length+3}'] = 'Manager Total'
            worksheet[f'L{length+4}'] = 500

            worksheet[f'N{length+4}'] = f'=SUM({provider_total},{commission_value},{gratuity_value},L{length+4})'


        wb.save(f'{destination}Master Payroll {date}.xlsx')

        #Cover Sheet
        row = str(len(medical_professionals) + len(massage_therapists) + n + 4)
        wb = openpyxl.load_workbook(f'{destination}Master Payroll {date}.xlsx')
        worksheet = wb['Cover Sheet']

        reg_pay_cell = f'D{row}'
        worksheet[reg_pay_cell] = f"='{name}'!K{length+4}"

        retail_commission = f'G{row}'
        worksheet[retail_commission] = f"='{name}'!K{length+7}"

        gratuity_cell = f'H{row}'
        worksheet[gratuity_cell] = f"='{name}'!K{length+10}"

        total_cell = f'K{row}'
        worksheet[total_cell] = f'=SUM(D{row},E{row},F{row},G{row},H{row})'

        wb.save(f'{destination}Master Payroll {date}.xlsx')



    for n, name in enumerate(front_desk):
        wb = openpyxl.load_workbook(f'{destination}Master Payroll {date}.xlsx')
        worksheet = wb[name]

        df = products.loc[products['Employee'] == name]
        length = len(df)

        desk_dict = {
        'Ashley Listorti' : 26,
        'Julia Ingledue' : 15,
        'Karina Castro' : 15,
        'Leslie Dawkins' : 26,
        'Alyssa Jake' : 17
    }

        #Hour
        hour_cell = f'B{length+3}'
        worksheet[hour_cell] = 'Hours'

        hour_number = f'B{length+4}'
        worksheet[hour_number] = 0

        #Rate
        rate_cell = f'C{length+3}'
        worksheet[rate_cell] = 'Rate'

        rate_number = f'C{length+4}'
        worksheet[rate_number] = desk_dict[name]

        #Total
        total = f'D{length+3}'
        worksheet[total] = 'Total'

        total_value = f'D{length+4}'
        worksheet[total_value] = f'={hour_number}*{rate_number}'

        #Seller Total
        seller = f'J{length+3}'
        worksheet[seller] = 'Product Commission'

        seller_sum = f'J{length+4}'
        worksheet[seller_sum] = f'=SUM(J2:J{length+1})'

        wb.save(f'{destination}Master Payroll {date}.xlsx')

        #Cover Sheet
        row = str(len(medical_professionals) + len(massage_therapists) + len(ests) + n + 5)
        wb = openpyxl.load_workbook(f'{destination}Master Payroll {date}.xlsx')
        worksheet = wb['Cover Sheet']

        reg_pay = f'D{row}'
        worksheet[reg_pay] = f"='{name}'!D{length+4}"

        total_pay = f'K{row}'
        worksheet[total_pay] = f"=SUM(D{row},E{row},F{row},G{row},H{row})"

        product_commission = f'G{row}'
        worksheet[product_commission] = f"='{name}'!J{length+4}"

        wb.save(f'{destination}Master Payroll {date}.xlsx')


    #Cover Sheet Column Names
    wb = openpyxl.load_workbook(f'{destination}Master Payroll {date}.xlsx')
    worksheet = wb['Cover Sheet']

    worksheet['C1'] = 1099
    worksheet['D1'] = 'Reg Pay/Service'
    worksheet['E1'] = 'Salary'
    worksheet['F1'] = 'Bonus'
    worksheet['G1'] = 'Retail Commission'
    worksheet['H1'] = 'Gratuity'
    worksheet['K1'] = 'Total'

    total_cell_number = len(medical_professionals) + len(massage_therapists) + len(ests) + len(front_desk) + len(management) + 7
    total_cell = f'K{str(total_cell_number)}'
    worksheet[total_cell] = f'=SUM(K2:K{total_cell_number-2})'

    wb.save(f'{destination}Master Payroll {date}.xlsx')